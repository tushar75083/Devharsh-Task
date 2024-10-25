import openpyxl
import os
import subprocess

# File name for storing user data
file_name = 'users.xlsx'

# Create a new Excel file with headers if it doesn't exist
def create_excel_file():
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Users'
        # Adding headers
        sheet.append(['Name', 'Email', 'Phone Number'])
        workbook.save(file_name)

# Function to add a new user
def add_user():
    name = input("Enter Name: ")
    email = input("Enter Email: ")
    phone_number = input("Enter Phone Number: ")

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    # Append the new user's details
    sheet.append([name, email, phone_number])
    # Save and close the workbook
    workbook.save(file_name)
    workbook.close()
    print("User added successfully.\n")

# Function to display all users
def display_users():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    # Check if there are any users stored
    if sheet.max_row == 1:
        print("No users found.\n")
        workbook.close()
        return

    print("\nStored Users:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Name: {row[0]}, Email: {row[1]}, Phone Number: {row[2]}")
    print()  # For better formatting
    workbook.close()

# Function to open the Excel file using the default application
def open_excel_file():
    try:
        if os.name == 'nt':  # For Windows
            os.startfile(file_name)
        elif os.name == 'posix':  # For macOS or Linux
            subprocess.call(['open' if os.uname().sysname == 'Darwin' else 'xdg-open', file_name])
        print("Opening the Excel file...\n")
    except Exception as e:
        print(f"Could not open the Excel file: {e}\n")

# Main function to show menu and handle user choice
def main():
    create_excel_file()
    while True:
        print("Menu:")
        print("1. Add User")
        print("2. Display Users")
        print("3. Open Excel File")
        print("4. Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            add_user()
        elif choice == '2':
            display_users()
        elif choice == '3':
            open_excel_file()
        elif choice == '4':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please enter 1, 2, 3, or 4.\n")

# Run the main function
if __name__ == "__main__":
    main()
